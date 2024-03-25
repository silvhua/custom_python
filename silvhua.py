import pandas as pd
import pickle
from datetime import datetime
import json
try:
    import openpyxl
except:
    pass
from datetime import datetime
import os

def convert_windows_path(path):
    path = f'{path}/'.replace('\\','/')
    return path

def append_timestamp(string):
    """
    Append timestamp to filename string.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    return f'{string}_{timestamp}' 

def load_and_describe_csv(
        filename, path, subset=None, id_column=None, 
        logger=None, logging_level=logging.INFO, **kwargs
    ):
    """
    Load a CSV as a dataframe and list the dataframe's columns and data types.

    Parameters:
    - filename (str)
    - path (raw string): Use the format r'<path>'. If None, file is saved in the same directory.
    - kwargs: Additional arguments to pass to pd.read_csv
    """
    messages_list = []
    logger = create_function_logger('load_and_describe_csv', logger, level=logging_level)
    df = load_csv(filename, path, **kwargs)
    if type (id_column) == int:
        id_column = df.columns[id_column]
    if subset == None:
        subset = df.columns.tolist() 
    if id_column != None:
        subset.remove(id_column)
    duplicate_rows = return_duplicate_rows(
        df, subset=subset, id_column=id_column, logger=logger, logging_level=logging_level
        )
    messages_list.append(f'\tNumber of null records: {df[subset].isnull().all(axis=1).sum()}')
    logger.debug('\n'.join(messages_list))
    logger.debug(df.dtypes)
    return df

def save_excel(
    df, filename, path=None, sheet_name=None, append_version=False, index=False, wrapping=True, 
    col_width=None, freeze_at='B2', overwrite=False
    ):
    """
    Export dataframe to Excel.
    Parameters:
    - df: Dataframe variable name.
    - filename: Root of the filename.
    - path (raw string): Use the format r'<path>'. If None, file is saved in the same directory.
    - append_version (bool): If true, append date and time to end of filename.
    - index (bool): If true, save index.
    - wrapping (bool): If true, enable text wrapping in Excel.
    - col_width (dict): Dictionary specifying column widths. Keys are column indices, values are column widths.
    """
    sheet_name = sheet_name if sheet_name else filename
    if (check_sheet_existence(filename, path, sheet_name=sheet_name) == False) | (overwrite == True):
        if path:
            path = convert_windows_path(path)
        if append_version:
            filename += f"_{datetime.now().strftime('%Y-%m-%d_%H%M')}"
        filepath = path + filename + '.xlsx'
        
        mode = 'a' if os.path.exists(filepath) else 'w'

        with pd.ExcelWriter(filepath, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            # Set the position of the sheet to be the left-most tab
            writer.sheets[sheet_name].index = 0
            
            # Access the workbook and the sheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            worksheet.freeze_panes = freeze_at
            if wrapping:
                # Set the text wrapping for all cells in the sheet
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')
        
            if col_width:
                # Set column widths
                for col_idx, width in col_width.items():
                    if (type(col_idx) == str) : # If col_idx is Excel column index such as 'A' or 'AA'
                        if len(col_idx) > 3:  # If col_idx is a column name
                            col_idx = df.columns.get_loc(col_idx)
                    if type(col_idx) != str:  # If col_idx is an integer or float
                        col_idx = chr(65 + col_idx)
                    worksheet.column_dimensions[col_idx].width = width
            # Save the workbook
            workbook.save(filepath)

        print('File saved:', path + filename + '.xlsx')
        print('Time completed:', datetime.now())
    else:
        pass
    return df

def check_sheet_existence(filename, path, sheet_name):
    if path:
        path = convert_windows_path(path)
    file_path = os.path.join(path, filename + '.xlsx')
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            print(f"The file '{filename}.xlsx' exists in '{path}' and the sheet '{sheet_name}' exists.")
            return True
        else:
            print(f"The file '{filename}.xlsx' exists in '{path}' but the sheet '{sheet_name}' does not exist.")
    else:
        print(f"The file '{filename}.xlsx' does not exist in '{path}'.")
    return False

# 2022-10-27 17:02 Update the sampling function to avoid loading entire dataframe.
def load_csv(filename,filepath,column1_as_index=False,truncate=None, usecols=None, sep=',', **kwargs):
    """
    Load a csv file as a dataframe using specified file path copied from windows file explorer.
    Back slashes in file path will be converted to forward slashes.
    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).
    - colum1_as_index (bool): If true, take the first column as the index. 
        Useful when importing CSV files from previously exported dataframes.

    Returns: dataframe object.
    """
    filename = f'{filepath}/'.replace('\\','/')+filename
    df = pd.read_csv(filename, usecols=usecols, sep=sep, **kwargs)
    if column1_as_index==True:
        df.set_index(df.columns[0], inplace=True)
        df.index.name = None
    print('Dataframe shape: ',df.shape)
    print('DataFrame columns:', [col for col in df.columns])
    print('\tTime completed:', datetime.now())

    if truncate:
        return df.sample(n=truncate,random_state=0)
    else:
        return df

def save_csv(df,filename,path=None,append_version=False, index=False):
    """
    Export dataframe to CSV.
    Parameters:
    - df: Dataframe variable name.
    - filename: Root of the filename.
    - filepath (raw string): Use the format r'<path>'. If None, file is saved in same director.
    - append_version (bool): If true, append date and time to end of filename.
    - index (bool): If true, save index.
    """
    if path:
        path = f'{path}/'.replace('\\','/')
    if append_version == True:
        filename+=f"_{datetime.now().strftime('%Y-%m-%d_%H%M')}"
    df.to_csv(path+filename+'.csv', index=index)
    print('File saved: ',path+filename+'.csv')
    print('\tTime completed:', datetime.now())

def save_text(text, filename, path=None, append_version=False, ext='txt'):
    """
    Save a string to a text file.
    Parameters:
    - text: The string to be saved.
    - filename: Root of the filename.
    - path (raw string): Use the format r'<path>'. If None, file is saved in the same directory.
    - append_version (bool): If True, append date and time to the end of the filename.
    - ext (string): Extension to append (no need to include dot as it will be added)
    """
    if path:
        path = f'{path}/'.replace('\\','/')
    if append_version:
        filename += f"_{datetime.now().strftime('%Y-%m-%d_%H%M')}"
    if ext[0] != '.':
        ext = '.'+ext
    with open(path+filename+ext, 'w') as file:
        file.write(text)
    print('File saved: ', path+filename+ext)
    print('\tTime completed:', datetime.now())

def load_txt(filename, filepath, encoding='utf-8'):
    """
    Load a text file as a string using the specified file path copied from Windows file explorer.
    Backslashes in the file path will be converted to forward slashes.

    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).

    Returns: string object.
    """
    filename = f'{filepath}/'.replace('\\','/') + filename
    with open(filename, 'r', encoding=encoding) as file:
        text = file.read()
    return text

def load_excel(filename,filepath,column1_as_index=False,truncate=None, usecols=None, sep=','):
    """
    Load an excel file as a dataframe using specified file path copied from windows file explorer.
    Back slashes in file path will be converted to forward slashes.
    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).
    - colum1_as_index (bool): If true, take the first column as the index. 
        Useful when importing CSV files from previously exported dataframes.

    Returns: dataframe object.
    """
    filename = f'{filepath}/'.replace('\\','/')+filename
    df = pd.read_excel(filename, usecols=usecols)
    if column1_as_index==True:
        df.set_index(df.columns[0], inplace=True)
        df.index.name = None
    print('Dataframe shape: ',df.shape)
    print('DataFrame columns:', [col for col in df.columns])
    print('\tTime completed:', datetime.now())

    if truncate:
        return df.sample(n=truncate,random_state=0)
    else:
        return df
def savepickle(model,filename, ext='sav', path=None,append_version=False):
    """
    Export object as a pickle.
    Parameters:
    - model: Model variable name.
    - filename: Root of the filename.
    - extension: Extension to append (do not include dot as it will be added)
    - filepath (raw string): Use the format r'<path>'. If None, file is saved in same director.
    - append_version (bool): If true, append date and time to end of filename.
    """
    if path:
        path = f'{path}/'.replace('\\','/')
    if append_version == True:
        filename+=datetime.now().strftime('%Y-%m-%d_%H%M')
    full_path = path+filename+'.'+ext if ext else path+filename
    with open (full_path, 'wb') as fh:
        pickle.dump(model, fh)
    print('File saved: ',full_path)
    print('\tTime completed:', datetime.now())

def loadpickle(filename,filepath):
    """
    Load a pickled model using specified file path copied from windows file explorer.
    Back slashes in file path will be converted to forward slashes.
    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).
    
    Returns saved object.
    """
    filename = f'{filepath}/'.replace('\\','/')+filename
    object = pickle.load(open(filename, 'rb'))
    print('\tTime completed:', datetime.now())
    if type(object) == pd.core.frame.DataFrame:
        print('Dataframe shape: ',object.shape)
        print('DataFrame columns:', [col for col in object.columns])
    if type(object) == dict:
        print('Dictionary keys:', [key for key in object.keys()])
    return object

def joblib_save(model,filename,path=None,append_version=False):
    """
    Export object with joblib.
    Parameters:
    - model: Model variable name.
    - filename: Root of the filename.
    - filepath (raw string): Use the format r'<path>'. If None, file is saved in same director.
    - append_version (bool): If true, append date and time to end of filename.
    """
    if path:
        path = f'{path}/'.replace('\\','/')
    if append_version == True:
        filename+=datetime.now().strftime('%Y-%m-%d_%H%M')
    with open (path+filename, 'wb') as fh:
        joblib.dump(model, fh)
    print('File saved: ',path+filename)
    print('\tTime completed:', datetime.now())


def joblib_load(filename,filepath):
    """
    Load a pickled model using specified file path copied from windows file explorer.
    Back slashes in file path will be converted to forward slashes.
    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).
    
    Returns saved object.
    """
    filename = f'{filepath}/'.replace('\\','/')+filename
    loaded_model = joblib.load(open(filename, 'rb'))
    return loaded_model

def save_output(df, filename=None, description=None, append_version=True, iteration_id=None, index=False,
    csv_path=r'C:\Users\silvh\OneDrive\lighthouse\Ginkgo coding\content-summarization\output\CSV',
    pickle_path=r'C:\Users\silvh\OneDrive\lighthouse\Ginkgo coding\content-summarization\output\pickles'
    ):
    """
    Save an Python object as both pickle and CSV. Automatically create filename using date and time 
    if not provided.
    
    """
    if description:
        filename = f'{description}_{datetime.now().strftime("%Y-%m-%d_%H%M")}'
        append_version = False
    elif filename == None:
        filename = f'{datetime.now().strftime("%Y-%m-%d_%H%M")}_outputs'
        append_version = False
    if iteration_id:
        filename += f'_{"{:02d}".format(iteration_id)}'
    try:
        savepickle(df, filename=filename, path=pickle_path, append_version=append_version)
        print('\tObject saved as pickle')
    except:
        print('Unable to save pickle')
    if (type(df) == pd.core.frame.DataFrame) & (csv_path != None):
        save_csv(df, filename=filename, path=csv_path, append_version=append_version, index=index)
        print('\tDataFrame saved as CSV')
    elif (type(df) == dict) & (csv_path != None):
        try:
            save_csv(pd.DataFrame(df), filename=filename, path=csv_path, append_version=append_version)
            print('\tDictionary converted to CSV')
        except:
            print('\tUnable to save CSV')


def save_to_json(obj, filename=None, description='output_dictionary', append_version=False,
    path=r'C:\Users\silvh\OneDrive\lighthouse\Ginkgo coding\content-summarization\output\json'
    ):
    """
    Save Python object as a JSON file.
    Parameters:
    - obj: Python object to be saved.
    - filename: Root of the filename.
    - path (raw string): Use the format r'<path>'. If None, file is saved in same directory as script.
    - append_version (bool): If true, append date and time to end of filename.
    """
    if description:
        filename = f'{description}_{datetime.now().strftime("%Y-%m-%d_%H%M")}'
        append_version = False
    elif filename == None:
        filename = f'{datetime.now().strftime("%Y-%m-%d_%H%M")}_outputs'
        append_version = False
    if path:
        path = f'{path}/'.replace('\\','/')
    if append_version:
        filename += f'_{datetime.now().strftime("%Y-%m-%d_%H%M%S")}'
    filename += '.json'
    with open(path+filename, 'w') as f:
        json.dump(obj, f, indent=4)
    print(f'Object saved as JSON: {path}{filename}')

def save_json_string(json_string, **kwargs):
    """
    Save a json string to a JSON file:
    1. Replaces single quotes with double quotes in the given JSON string and converts it to a dictionary.
    2. Using the `save_to_json` function with the `json_string` argument, saves the dictionary to a JSON file 
    with the provided keyword arguments.
    """
    json_string = json_string.replace("'", '"')
    dictionary = json.loads(json_string)
    print('JSON string converted to dictionary.')
    save_to_json(dictionary, **kwargs)

def load_json(filename, filepath):
    """
    Load a JSON file using specified file path copied from windows file explorer.
    Back slashes in file path will be converted to forward slashes.

    Arguments:
    - filepath (raw string): Use the format r'<path>'.
    - filename (string).
    """
    filename = f'{filepath}/'.replace('\\','/')+filename
    with open(filename) as file:
        return json.load(file)
    
def dict_list_to_json(dict_list, primary_key='id'):
    """
    Converts a list of dictionaries to a JSON object, using a specified primary key.

    Parameters:
        dict_list (list): A list of dictionaries to be converted.
        primary_key (str, optional): The key to be used as the primary key in the resulting JSON object. Defaults to 'id'.

    Returns:
        dict: A JSON object where each dictionary in the input list is a value, and the value of the primary key is the key.
    """
    new_dict_list = []
    result = dict()
    for dictionary in dict_list:
        new_dict = {}
        for key, value in dictionary.items():
            key = key.replace(' ', '_').replace('\n','')
            value = value.replace(' ', '_').replace('\n','') if type(value) == str else value
            new_dict[key] = value
        new_dict_list.append(new_dict)
    if primary_key:
        for dictionary in new_dict_list:
            key_value = dictionary[primary_key]
            # print(f'key value: {key_value}\n\n')
            result[key_value] = dictionary
    else:
        for index, dictionary in enumerate(new_dict_list):
            result[index] = dictionary
    return result

def list_files(path, substring='.csv'):
    """
    List the filenames in a given filepath.

    Parameters:
    - path (str): Filepath.
    - substring (str): Return only filenames containing this string/extension. 
        Default is '.csv'.

    Returns: 
    - filenames_list (list)
    """
    filenames_list = [file for file in os.listdir(path) if substring in file]
    print(f'Number of files {f" containing `{substring}`" if substring else ""}: {len(filenames_list)}')
    return filenames_list

def sample_csv_files(path, nrows=5, **kwargs):
    """
    Load a few rows from all CSV files from a given path and substring.

    Returns:
    - df_dict (dict): A dictionary of dataframes where the key is the filename without the extension.
    """
    filenames_list = list_files(path, substring='.csv')
    df_dict = dict()
    for file in filenames_list:
        dict_key = file.split('.')[0]
        print(f'**{file}**:')
        df_dict[dict_key] = load_csv(file, path, nrows=nrows, **kwargs)
        print()

    return df_dict

    
# def time_columns(df,time_column,format='%H%M'):
#     """ 
#     Take the time in a dateframes to create new columns with datetime objects.

#     Parmaters:
#     - df: Dataframe.
#     - time_column (string or list of strings): Name of the column containing the time.
#     - Format: Original time format in the dateframe. Default is in flight format: '%H%M'
    
#     Make sure to do the following import: 
#     from datetime import datetime
#     """
#     if type(time_column) == str:
#         time_column = [time_column]
#     for column in time_column:
#         df[str(column+'_time')] = pd.to_datetime(df[column],format=format)
   
#     return df
